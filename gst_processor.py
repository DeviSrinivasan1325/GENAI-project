"""
GST Invoice Processing Logic
Extracted from GST_documentation_version4.ipynb
"""

import os
import re
import json
import time
import pytesseract
import pdfplumber
from PIL import Image, ImageEnhance, ImageFilter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import google.generativeai as genai

# ─── Configuration ───────────────────────────────────────────────────────────
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
SUPPORTED_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}

EXTRACTION_PROMPT = """
You are a GST invoice data extraction engine for Indian businesses.

Return STRICTLY valid JSON only — no markdown, no backticks, no explanation.

{{
  "vendor_name": "...",
  "gstin": "...",
  "invoice_number": "...",
  "invoice_date": "DD-MM-YYYY",
  "taxable_amount": <number>,
  "cgst": <number>,
  "sgst": <number>,
  "igst": <number>,
  "total_amount": <number>
}}

RULES — read carefully:

vendor_name:
  - The SELLER company name. Usually the TOPMOST bold/all-caps heading.
  - Do NOT extract: subtitles, service descriptions, taglines, project details.
  - Examples of WRONG values: "Software Development & IT Consulting", "Premium Catering & Event Management"
  - Examples of CORRECT values: "NEXATECH SOLUTIONS LLP", "ROYAL RASOI CATERERS"

gstin:
  - SELLER GSTIN only (appears near seller address, NOT Bill To / Consignee).
  - Exactly 15 characters. Return null if not found.
  - Common OCR errors to watch: Z may appear as 2, O may appear as 0.

taxable_amount:
  - Sum of all line items BEFORE tax.
  - Indian format: "4,20,000" = 420000. Remove ALL commas.

total_amount:
  - The GROSS total BEFORE any advance deduction.
  - Look for: GRAND TOTAL / TOTAL AMOUNT / TOTAL INVOICE VALUE / TOTAL AMOUNT PAYABLE.
  - If invoice shows "Balance Payable" after deducting advance, look for the NOTE line
    that says "grand total of Rs. X" — use X, not the balance.
  - Example: Catering invoice shows Balance Payable Rs.3,64,750 but
    note says grand total Rs.4,64,750 → use 4,64,750.

cgst / sgst / igst:
  - Set 0 for any tax type not present.
  - CGST + SGST = intra-state. IGST = inter-state. Never both.

All numeric fields must be plain numbers — NOT strings.

Invoice Text:
{text}
"""

GSTIN_PATTERN = r"^\d{2}[A-Z]{5}\d{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$"

HEADERS = ["S.No", "Vendor Name", "GSTIN", "Invoice No.", "Invoice Date",
           "Taxable (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Total (₹)",
           "Confidence", "Processed On"]
WIDTHS  = [6, 28, 20, 18, 14, 16, 12, 12, 12, 16, 11, 18]
CURRENCY_COLS = {6, 7, 8, 9, 10}

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
TOTAL_FILL   = PatternFill("solid", start_color="D6E4F0")
ALT_FILL     = PatternFill("solid", start_color="F2F7FB")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
WARN_FILL    = PatternFill("solid", start_color="FFF3CD")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TOTAL_FONT   = Font(bold=True, name="Arial", size=10)
DATA_FONT    = Font(name="Arial", size=9)
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY_FMT = '#,##0.00'
PCT_FMT      = '0%'
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")
RIGHT        = Alignment(horizontal="right",  vertical="center")

TO_DIGIT  = {'O': '0', 'I': '1', 'S': '5', 'B': '8', 'G': '6'}
TO_LETTER = {'0': 'O', '1': 'I', '5': 'S', '8': 'B'}


# ─── Text Extraction ──────────────────────────────────────────────────────────

def preprocess_image(image):
    image = image.convert("L")
    image = ImageEnhance.Contrast(image).enhance(2.0)
    image = image.filter(ImageFilter.SHARPEN)
    return image


def table_to_text(table):
    lines = []
    for row in table:
        cells = [str(c).strip() if c else "" for c in row]
        non_empty = [c for c in cells if c]
        if not non_empty:
            continue
        if len(cells) == 2:
            for c in cells:
                if c:
                    lines.append(c)
        elif len(cells) == 4 and any(":" in c for c in cells):
            for i in range(0, len(cells), 2):
                k = cells[i].strip()
                v = cells[i + 1].strip() if i + 1 < len(cells) else ""
                if k or v:
                    lines.append(f"{k} {v}".strip())
        else:
            lines.append("  |  ".join(non_empty))
    return "\n".join(lines)


def extract_text_from_pdf(file_path):
    parts = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            raw_text  = page.extract_text() or ""
            tables    = page.extract_tables()
            table_texts = [table_to_text(t) for t in tables if t]
            if not raw_text.strip() and not table_texts:
                pil = page.to_image(resolution=200).original  # 200 DPI sufficient for OCR, faster than 300
                parts.append(pytesseract.image_to_string(
                    preprocess_image(pil), config="--psm 6 --oem 3"))
                continue
            page_text = raw_text + "\n"
            if table_texts:
                page_text += "\n--- STRUCTURED TABLE DATA ---\n" + "\n\n".join(table_texts)
            parts.append(page_text)
    return "\n\n".join(parts).strip()


def extract_text(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: '{ext}'")
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    return pytesseract.image_to_string(
        preprocess_image(Image.open(file_path)), config="--psm 6 --oem 3"
    ).strip()


# ─── OCR Correction ───────────────────────────────────────────────────────────

def correct_gstin_ocr(gstin):
    if not gstin:
        return gstin
    g = list(str(gstin).strip().upper())
    if len(g) != 15:
        return ''.join(g)
    for i in [0, 1, 7, 8, 9, 10]:
        if g[i] in TO_DIGIT:
            g[i] = TO_DIGIT[g[i]]
    for i in [2, 3, 4, 5, 6, 11]:
        if g[i] in TO_LETTER:
            g[i] = TO_LETTER[g[i]]
    if g[13] != 'Z':
        g[13] = 'Z'
    return ''.join(g)


# ─── LLM Extraction ───────────────────────────────────────────────────────────

def parse_indian_number(value):
    if value is None:
        return 0.0
    s = re.sub(r"[^\d.]", "", str(value).replace(",", ""))
    try:
        return float(s) if s else 0.0
    except ValueError:
        return 0.0


# Module-level model cache — initialized once per session, not per invoice
_gemini_model = None

def _get_model(api_key):
    """Return cached Gemini model, initializing only if api_key changed."""
    global _gemini_model, _gemini_api_key
    if _gemini_model is None or getattr(_get_model, "_last_key", None) != api_key:
        genai.configure(api_key=api_key)
        _gemini_model = genai.GenerativeModel("gemini-2.5-flash")
        _get_model._last_key = api_key
    return _gemini_model


def extract_invoice_data(text, api_key, max_retries=3):
    if not text or len(text.strip()) < 10:
        return None, "Text too short for extraction."
    model = _get_model(api_key)  # cached — no re-init overhead
    for attempt in range(1, max_retries + 1):
        try:
            response = model.generate_content(EXTRACTION_PROMPT.format(text=text))
            raw = response.text or response.candidates[0].content.parts[0].text
            cleaned = raw.strip()
            if "```" in cleaned:
                for part in cleaned.split("```"):
                    s = part.strip().lstrip("json").strip()
                    if s.startswith("{"):
                        cleaned = s
                        break
            data = json.loads(cleaned.strip())
            for f in ["taxable_amount", "cgst", "sgst", "igst", "total_amount"]:
                data[f] = parse_indian_number(data.get(f))
            return data, None
        except Exception as e:
            err = str(e)
            if "429" in err or "quota" in err.lower():
                m = re.search(r'seconds.*?(\d+)', err)
                wait = int(m.group(1)) + 5 if m else 15 * attempt
                if attempt < max_retries:
                    time.sleep(wait)
                    continue
                return None, "Gemini rate limit — all retries exhausted."
            elif isinstance(e, json.JSONDecodeError):
                if attempt < max_retries:
                    time.sleep(3)
                    continue
                return None, "LLM returned invalid JSON."
            else:
                return None, f"LLM error: {e}"
    return None, "Extraction failed after retries."


# ─── Rule-Based Extraction ────────────────────────────────────────────────────

def extract_amounts_from_text(text):
    def find_amount(patterns, text):
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                raw = m.group(1).replace(',', '')
                try:
                    return float(raw)
                except:
                    pass
        return None

    result = {}
    result['taxable_amount'] = find_amount([
        r'taxable\s+amount[^\d]*([\d,]+\.?\d*)',
        r'taxable[^\d]*([\d,]+\.?\d*)',
    ], text)
    result['cgst'] = find_amount([
        r'cgst\s*@?\s*[\d.]+%?[:\s]*(?:rs\.?\s*)?([\d,]+\.?\d*)',
        r'cgst[:\s]+(?:rs\.?\s*)?([\d,]+\.?\d*)',
    ], text)
    result['sgst'] = find_amount([
        r'sgst\s*@?\s*[\d.]+%?[:\s]*(?:rs\.?\s*)?([\d,]+\.?\d*)',
        r'sgst[:\s]+(?:rs\.?\s*)?([\d,]+\.?\d*)',
    ], text)
    result['igst'] = find_amount([
        r'igst\s*@?\s*[\d.]+%?[:\s]*(?:rs\.?\s*)?([\d,]+\.?\d*)',
        r'igst[:\s]+(?:rs\.?\s*)?([\d,]+\.?\d*)',
    ], text)
    gross_note = re.search(
        r'grand\s+total\s+of\s+(?:rs\.?\s*)?([\d,]+\.?\d*)', text, re.IGNORECASE)
    if gross_note:
        result['total_amount'] = float(gross_note.group(1).replace(',', ''))
    else:
        result['total_amount'] = find_amount([
            r'total\s+invoice\s+value[:\s]*(?:rs\.?\s*)?([\d,]+\.?\d*)',
            r'total\s+amount\s+payable[:\s]*(?:rs\.?\s*)?([\d,]+\.?\d*)',
            r'grand\s+total[:\s]*(?:rs\.?\s*)?([\d,]+\.?\d*)',
            r'total\s+amount[:\s]*(?:rs\.?\s*)?([\d,]+\.?\d*)',
        ], text)
    return {k: v for k, v in result.items() if v is not None}


def merge_rule_llm(rule_data, llm_data):
    if not llm_data:
        return llm_data
    merged = dict(llm_data)
    for field in ['taxable_amount', 'cgst', 'sgst', 'igst', 'total_amount']:
        if field in rule_data and rule_data[field] is not None:
            merged[field] = rule_data[field]
    return merged


# ─── Validation ───────────────────────────────────────────────────────────────

def validate_gstin(gstin):
    if not gstin:
        return None
    g = str(gstin).strip().upper()
    if len(g) != 15:
        return None
    return g if re.match(GSTIN_PATTERN, g) else None


def validate_date(date_str):
    if not date_str:
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(str(date_str).strip(), fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return date_str


def validate_tax_math(data):
    taxable  = data.get("taxable_amount", 0) or 0
    computed = round(taxable + (data.get("cgst", 0) or 0) +
                     (data.get("sgst", 0) or 0) + (data.get("igst", 0) or 0), 2)
    actual   = round(data.get("total_amount", 0) or 0, 2)
    return abs(computed - actual) <= 2.0, computed, actual


def _load_existing_df(excel_file):
    """Load, clean and normalize existing invoices from Excel into a DataFrame.
    Returns empty DataFrame if file doesn't exist yet.
    Called ONCE per invoice in process_invoice — result is reused for both
    duplicate check and save, eliminating the previous double-read."""
    if not os.path.exists(excel_file):
        return pd.DataFrame()
    try:
        df = pd.read_excel(excel_file, sheet_name="Invoices", header=2)
        df.columns = df.columns.str.strip()
        if "Vendor Name" not in df.columns:
            return pd.DataFrame()
        df = df[
            df["Vendor Name"].notna() &
            (df["Vendor Name"].astype(str).str.strip() != "") &
            (df["S.No"].astype(str).str.strip() != "TOTAL")
        ]
        return df.rename(columns={
            "Vendor Name": "vendor_name", "GSTIN": "gstin",
            "Invoice No.": "invoice_number", "Invoice Date": "invoice_date",
            "Taxable (₹)": "taxable_amount", "CGST (₹)": "cgst",
            "SGST (₹)": "sgst", "IGST (₹)": "igst",
            "Total (₹)": "total_amount", "Confidence": "confidence",
            "Processed On": "processed_on",
        }).drop(columns=["S.No"], errors="ignore")
    except Exception:
        return pd.DataFrame()


def check_duplicate(data, excel_file, existing_df=None):
    """Check for duplicates using a pre-loaded DataFrame when available,
    avoiding an extra Excel read when called from process_invoice."""
    df = existing_df if existing_df is not None else _load_existing_df(excel_file)
    if df.empty:
        return False
    inv    = str(data.get("invoice_number", "")).strip().lower()
    vendor = str(data.get("vendor_name", "")).strip().lower()
    gstin  = str(data.get("gstin", "")).strip().upper()
    for _, row in df.iterrows():
        existing_inv    = str(row.get("invoice_number", "")).strip().lower()
        existing_vendor = str(row.get("vendor_name", "")).strip().lower()
        existing_gstin  = str(row.get("gstin", "")).strip().upper()
        # Primary: invoice number + GSTIN — unique in India's GST system
        if inv and gstin and existing_inv == inv and existing_gstin == gstin:
            return True
        # Fallback: fuzzy vendor prefix match — catches OCR drift ("LLP" dropped etc.)
        if existing_inv == inv and vendor and existing_vendor:
            short = min(vendor, existing_vendor, key=len)
            long_ = max(vendor, existing_vendor, key=len)
            if short and long_.startswith(short):
                return True
    return False


def compute_confidence(data, gstin_was_corrected=False, date_parsed=True, tax_math_ok=True):
    score = 1.0
    if not data.get("gstin"):       score -= 0.20
    elif gstin_was_corrected:       score -= 0.10
    if not date_parsed:             score -= 0.10
    if not tax_math_ok:             score -= 0.20
    if not data.get("invoice_number"): score -= 0.15
    if not data.get("vendor_name"):    score -= 0.15
    return round(max(score, 0.0), 2)


def validate_data(data, excel_file, existing_df=None):
    result = {"data": data, "warnings": [], "errors": [],
              "is_duplicate": False, "confidence": 1.0}
    if not data:
        result["errors"].append("No data extracted.")
        return result

    raw_gstin         = data.get("gstin", "")
    corrected         = correct_gstin_ocr(raw_gstin) if raw_gstin else raw_gstin
    gstin_was_corrected = (corrected != raw_gstin) and bool(raw_gstin)
    if gstin_was_corrected:
        result["warnings"].append(f"GSTIN OCR corrected: '{raw_gstin}' → '{corrected}'")
    valid_gstin = validate_gstin(corrected)
    if corrected and not valid_gstin:
        result["warnings"].append(
            f"GSTIN still invalid after correction: '{corrected}' — saved as null.")
    data["gstin"] = valid_gstin

    data["invoice_date"] = validate_date(data.get("invoice_date"))
    date_parsed = bool(data["invoice_date"])
    if not date_parsed:
        result["warnings"].append("Invoice date could not be parsed.")

    ok, computed, actual = validate_tax_math(data)
    if not ok:
        result["warnings"].append(
            f"Tax math mismatch: ₹{computed:,.2f} ≠ ₹{actual:,.2f} "
            f"(diff ₹{abs(computed - actual):.2f}) — verify manually.")

    for f in ["vendor_name", "invoice_number", "taxable_amount", "total_amount"]:
        if not data.get(f):
            result["errors"].append(f"Missing required field: '{f}'")

    if check_duplicate(data, excel_file, existing_df=existing_df):
        result["is_duplicate"] = True
        result["errors"].append(
            f"DUPLICATE: Invoice '{data.get('invoice_number')}' from "
            f"'{data.get('vendor_name')}' already exists.")

    result["confidence"] = compute_confidence(data, gstin_was_corrected, date_parsed, ok)
    result["data"] = data
    return result


# ─── Excel Output ─────────────────────────────────────────────────────────────

def _sc(cell, font=None, fill=None, alignment=None, border=None, num_format=None):
    if font:       cell.font          = font
    if fill:       cell.fill          = fill
    if alignment:  cell.alignment     = alignment
    if border:     cell.border        = border
    if num_format: cell.number_format = num_format


def _write_invoices_sheet(ws, df):
    ws.merge_cells("A1:L1")
    ws["A1"].value = "GST Invoice Register"
    _sc(ws["A1"], Font(bold=True, color="1F4E79", name="Arial", size=14),
        PatternFill("solid", start_color="D6E4F0"), CENTER, BORDER)
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:L2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    _sc(ws["A2"], Font(italic=True, color="595959", name="Arial", size=9), alignment=CENTER)
    ws.row_dimensions[2].height = 14
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=3, column=col, value=h)
        _sc(cell, HEADER_FONT, HEADER_FILL, CENTER, BORDER)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20
    data_start = 4
    for i, (_, row) in enumerate(df.iterrows()):
        r = data_start + i
        conf = row.get("confidence", 1.0)
        fill = WARN_FILL if (isinstance(conf, float) and conf < 0.7) else \
               (ALT_FILL if i % 2 == 1 else WHITE_FILL)
        vals = [i + 1, row.get("vendor_name"), row.get("gstin"),
                row.get("invoice_number"), row.get("invoice_date"),
                row.get("taxable_amount"), row.get("cgst"),
                row.get("sgst"), row.get("igst"),
                row.get("total_amount"), row.get("confidence"),
                row.get("processed_on")]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            align = CENTER if col in {1, 11} else (RIGHT if col in CURRENCY_COLS else LEFT)
            fmt   = (CURRENCY_FMT if col in CURRENCY_COLS else
                     PCT_FMT if col == 11 else None)
            _sc(cell, DATA_FONT, fill, align, BORDER, fmt)
        ws.row_dimensions[r].height = 17
    data_end   = data_start + len(df) - 1
    totals_row = data_end + 1
    col_map    = {6: "F", 7: "G", 8: "H", 9: "I", 10: "J"}
    for col in range(1, 13):
        cell = ws.cell(row=totals_row, column=col)
        _sc(cell, TOTAL_FONT, TOTAL_FILL, border=BORDER)
        if col == 1:
            cell.value = "TOTAL"; cell.alignment = CENTER
        elif col in col_map:
            L = col_map[col]
            cell.value = f"=SUM({L}{data_start}:{L}{data_end})"
            cell.alignment = RIGHT; cell.number_format = CURRENCY_FMT
        else:
            cell.alignment = LEFT
    ws.row_dimensions[totals_row].height = 20
    ws.freeze_panes = "A4"


def _write_summary_sheet(ws, df):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:B1")
    ws["A1"].value = "GST Summary Report"
    _sc(ws["A1"], Font(bold=True, color="1F4E79", name="Arial", size=13),
        PatternFill("solid", start_color="D6E4F0"), CENTER, BORDER)
    ws.row_dimensions[1].height = 26
    rows = [
        ("Total Invoices Processed", len(df)),
        ("Total Taxable Amount",     df["taxable_amount"].sum() if "taxable_amount" in df else 0),
        ("Total CGST",               df["cgst"].sum()           if "cgst" in df else 0),
        ("Total SGST",               df["sgst"].sum()           if "sgst" in df else 0),
        ("Total IGST",               df["igst"].sum()           if "igst" in df else 0),
        ("", ""),
        ("FINAL AMOUNT PAYABLE",     df["total_amount"].sum()   if "total_amount" in df else 0),
    ]
    for r_idx, (label, value) in enumerate(rows, start=2):
        a = ws.cell(row=r_idx, column=1, value=label)
        b = ws.cell(row=r_idx, column=2, value=value)
        is_final = label == "FINAL AMOUNT PAYABLE"
        is_blank = label == ""
        fill = (PatternFill("solid", start_color="1F4E79") if is_final else
                PatternFill("solid", start_color="F2F7FB") if not is_blank else WHITE_FILL)
        fc = "FFFFFF" if is_final else "000000"
        for cell in [a, b]:
            _sc(cell, Font(bold=is_final, name="Arial", size=11 if is_final else 10, color=fc),
                fill, LEFT, BORDER if not is_blank else None)
        if isinstance(value, (int, float)) and not is_blank:
            b.number_format = CURRENCY_FMT
            b.alignment = RIGHT
        ws.row_dimensions[r_idx].height = 22 if is_final else 18


def save_to_excel(record, file_name, existing_df=None):
    """Append record to Excel. Accepts pre-loaded existing_df to avoid
    re-reading the file when called from process_invoice."""
    rec = {k: record.get(k) for k in
           ["vendor_name", "gstin", "invoice_number", "invoice_date",
            "taxable_amount", "cgst", "sgst", "igst", "total_amount", "confidence"]}
    rec["processed_on"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    # Use pre-loaded df if provided — avoids a second pd.read_excel per invoice
    df = existing_df if existing_df is not None else _load_existing_df(file_name)
    df = pd.concat([df, pd.DataFrame([rec])], ignore_index=True)
    wb = Workbook()
    ws_inv = wb.active
    ws_inv.title = "Invoices"
    _write_invoices_sheet(ws_inv, df)
    _write_summary_sheet(wb.create_sheet("Summary"), df)
    wb.save(file_name)


def get_summary(file_name):
    """Return summary dict from the Excel file."""
    if not os.path.exists(file_name):
        return None
    df = pd.read_excel(file_name, sheet_name="Invoices", header=2)
    df.columns = df.columns.str.strip()
    df = df[
        df["Vendor Name"].notna() &
        (df["Vendor Name"].astype(str).str.strip() != "") &
        (df["S.No"].astype(str).str.strip() != "TOTAL")
    ]
    return {
        "count":   int(len(df)),
        "taxable": float(df["Taxable (₹)"].sum()) if "Taxable (₹)" in df else 0,
        "cgst":    float(df["CGST (₹)"].sum())    if "CGST (₹)"    in df else 0,
        "sgst":    float(df["SGST (₹)"].sum())    if "SGST (₹)"    in df else 0,
        "igst":    float(df["IGST (₹)"].sum())    if "IGST (₹)"    in df else 0,
        "total":   float(df["Total (₹)"].sum())   if "Total (₹)"   in df else 0,
    }


# ─── Full Pipeline ────────────────────────────────────────────────────────────

def process_invoice(file_path, api_key, excel_file):
    """
    Run the full pipeline on a single invoice file.
    Excel is read ONCE at the start and reused for duplicate check + save.
    Returns a dict with keys: success, data, warnings, errors, confidence, skipped_reason
    """
    result = {
        "success": False,
        "data": {},
        "warnings": [],
        "errors": [],
        "confidence": 0,
        "skipped_reason": None,
    }

    # Load existing invoices ONCE — reused for duplicate check and save
    existing_df = _load_existing_df(excel_file)

    # Step 1 — OCR / text extraction
    try:
        text = extract_text(file_path)
    except Exception as e:
        result["errors"].append(f"Text extraction failed: {e}")
        return result

    if not text or len(text.strip()) < 10:
        result["errors"].append("Could not extract readable text from this file.")
        return result

    # Step 2 — Rule-based extraction
    rule_data = extract_amounts_from_text(text)

    # Step 3 — LLM extraction (model is cached — no re-init overhead)
    llm_data, llm_err = extract_invoice_data(text, api_key)
    if llm_err and not llm_data:
        result["errors"].append(llm_err)
        return result

    # Step 4 — Merge
    data = merge_rule_llm(rule_data, llm_data)
    if not data:
        result["errors"].append("Extraction returned no data.")
        return result

    # Step 5 — Validate (passes existing_df — no extra Excel read)
    vr = validate_data(data, excel_file, existing_df=existing_df)
    result["warnings"] = vr["warnings"]
    result["errors"]   = vr["errors"]
    result["confidence"] = vr["confidence"]
    result["data"]       = vr["data"]

    if vr["is_duplicate"]:
        result["skipped_reason"] = "duplicate"
        return result

    if vr["errors"]:
        result["skipped_reason"] = "validation_failed"
        return result

    # Step 6 — Save (passes existing_df — no extra Excel read)
    vr["data"]["confidence"] = vr["confidence"]
    save_to_excel(vr["data"], excel_file, existing_df=existing_df)
    result["success"] = True
    return result
